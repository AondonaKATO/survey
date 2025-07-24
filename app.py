from flask import Flask, request, jsonify, render_template, send_file
import openpyxl
import os
from datetime import datetime

app = Flask(__name__)

FILE_PATH = 'survey_responses.xlsx'

def initialize_excel():
    if not os.path.exists(FILE_PATH):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Survey Responses"
        headers = [
            "Timestamp", "Participation", "Preferences", "Prompts", "Themes",
            "Friday Fun", "Format", "Tone", "Academic Tasks",
            "Literary", "Additional Suggestions"
        ]
        ws.append(headers)
        wb.save(FILE_PATH)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data received"}), 400

    try:
        wb = openpyxl.load_workbook(FILE_PATH)
        ws = wb.active

        row = [
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            data.get("participation", ""),
            data.get("preferences", ""),
            data.get("prompts", ""),
            data.get("themes", ""),
            data.get("friday_fun", ""),
            data.get("format", ""),
            data.get("tone", ""),
            data.get("academic_tasks", ""),
            data.get("literary", ""),
            data.get("additional", "")
        ]

        ws.append(row)
        wb.save(FILE_PATH)

        return jsonify({"message": "Response recorded successfully."}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/admin')
def admin():
    try:
        wb = openpyxl.load_workbook(FILE_PATH)
        sheet = wb.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        return render_template("admin.html", data=data)
    except Exception as e:
        return f"Error reading data: {e}"

@app.route('/download')
def download_excel():
    return send_file(FILE_PATH, as_attachment=True)

if __name__ == '__main__':
    initialize_excel()
    app.run(debug=True, host="0.0.0.0", port=10000)

